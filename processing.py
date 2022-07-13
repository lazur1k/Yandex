from openpyxl import load_workbook
import pandas as pd
import os

def processing(parser_type, main_dirrectory):
    ads_type = parser_type
    directoy = main_dirrectory

    tmp_excel = pd.DataFrame()
    for file in os.listdir(f'{directoy}/{ads_type}/excel/'):
        if file.endswith('.xlsx') and file != 'priceHistory.xlsx':
            db = pd.read_excel(f'{directoy}/{ads_type}/excel/{file}').iloc[:, 1:]
            tmp_excel = pd.concat([tmp_excel, db], ignore_index=True)
    tmp_excel = tmp_excel.drop_duplicates()
    
    price_db = pd.read_excel(f'{directoy}/{ads_type}/excel/priceHistory.xlsx').iloc[:, 1:]
    price_db = price_db.drop_duplicates()
    
    new_db = pd.merge(left = tmp_excel, right = price_db, how = 'left', on='url')
    new_db = new_db.dropna(subset=['id'])
    new_db = new_db.drop_duplicates(subset=['url'])        
    new_db.to_excel(f'{directoy}/{ads_type}/{ads_type}({len(new_db)}).xlsx')
    
    wb = load_workbook(f'{directoy}/{ads_type}/{ads_type}({len(new_db)}).xlsx')
    sheet = wb.active

    for cell in sheet.iter_rows(2,sheet.max_row, 4, 5):
        id = cell[0].value.split('/')[-2]
        sheet.cell(cell[0].row, 3).value = id

    wb.save(f'{directoy}/{ads_type}/{ads_type}({len(new_db)}).xlsx')
    wb.close()
    
    main_db = pd.read_excel(f'{directoy}/{ads_type}/{ads_type}({len(new_db)}).xlsx').iloc[:, 1:]
    scans = main_db['id'].values.tolist()
    for img in os.listdir(f'{directoy}/{ads_type}/screenshots'):
        if int(img.split('.')[0]) not in scans:
            os.remove(f'{directoy}/{ads_type}/sceenshots/{img}')
    
if __name__ == '__main__':
    processing('Квартиры', 'K:/Сектор мониторинга РН/Аналитика/1. СБОР ИНФОРМАЦИИ/Выгрузка - Яндекс/01.05.2022/')